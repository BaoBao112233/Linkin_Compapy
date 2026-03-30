# ==============================
# LinkedIn Company Scraper
# File: exporter.py
# Xuất dữ liệu ra CSV / JSON / Excel
# ==============================

import os
import json
from pathlib import Path
from typing import Iterable

import pandas as pd
from loguru import logger

import config


def ensure_output_dir():
    Path(config.OUTPUT_DIR).mkdir(parents=True, exist_ok=True)


def save_csv(records: list[dict], path: str = config.OUTPUT_CSV):
    ensure_output_dir()
    df = pd.DataFrame(records)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    logger.success(f"Đã lưu {len(records)} công ty → {path}")


def save_json(records: list[dict], path: str = config.OUTPUT_JSON):
    ensure_output_dir()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    logger.success(f"Đã lưu {len(records)} công ty → {path}")


def save_excel(records: list[dict], path: str = config.OUTPUT_XLSX):
    ensure_output_dir()
    df = pd.DataFrame(records)
    # Tô màu tiêu đề
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Companies")
        ws = writer.sheets["Companies"]
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0A66C2")  # LinkedIn blue
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Tự động co giãn độ rộng cột
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    logger.success(f"Đã lưu {len(records)} công ty → {path}")


def save_all(records: list[dict]):
    """Lưu cả 3 định dạng cùng lúc."""
    save_csv(records)
    save_json(records)
    save_excel(records)
